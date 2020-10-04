#!/usr/bin/env python
# Created by Filippo Pisello

from os import path, getcwd
import string

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

class Custom_excel:

    def __init__(self, dataframe, file_name, sheet_name="Sheet1",
                 keep_index=False, format_header=True, format_index_main=False,
                 format_index_light=True, color_main="0066cc", color_light="b2beb5",
                 font_color_main="ffffff", size_font_main=12, bold_font_main=True,
                 font_color_light="000000", size_font_light=11, bold_font_light=False,
                 font_size_body=11, alignment_main="center", alignment_light="left",
                 alignment_body="center", custom_width=20, check_file_name=True):
        self.df = dataframe
        self.ind_len, self.col_len = dataframe.shape
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.keep_index = keep_index
        self.check_file_name = check_file_name
        
        self.workbook_sheet_exist = []
        self.workbook = " "
        self.sheet = " "
        
        self.col_depth = self.ind_depth = " "
        self.header = self.index = self.body = []
        
        self.format_header = format_header
        self.format_index_main = format_index_main
        self.format_index_light =format_index_light
        self.main_color = color_main
        self.light_color = color_light
        self.main_font_color = font_color_main
        self.main_font_size = size_font_main
        self.main_font_bold = bold_font_main
        self.light_font_color = font_color_light
        self.light_font_size = size_font_light
        self.light_font_bold = bold_font_light
        self.body_font_size = font_size_body
        self.main_alignment = alignment_main
        self.light_alignment = alignment_light
        self.body_alignment = alignment_body
        self.custom_width = custom_width
    
    # -------------------------------------------------------------------------
    # 1) Main function running the complete process
    # -------------------------------------------------------------------------
    def to_custom_excel(self):
        # File saving process
        self.correct_file_name()
        self.workbook_sheet_exist, self.workbook = self.check_file_existence()
        self.save_df_to_excel()
        
        # Table dimensions analysis
        self.col_depth, self.ind_depth = self.find_index_columns_depth()
        header_coordinates = self.find_header_coordinates()
        index_coordinates = self.find_index_coordinates()
        body_coordinates = [[header_coordinates[0][0], index_coordinates[0][1]],
                            [header_coordinates[1][0], index_coordinates[1][1]]]
        self.header = self.rectangle_of_cells(header_coordinates)
        self.index = self.rectangle_of_cells(index_coordinates)
        self.body = self.rectangle_of_cells(body_coordinates)
        
        # Table customization
        self.workbook, self.sheet = self.get_workbook_sheet()
        if self.format_header:
            self.format_as_main(self.header)
        if self.format_index_light:
            self.format_as_light(self.index)
        if self.format_index_main:
            self.format_as_main(self.index)
        # If index was not formatted, format it as the body
        self.format_body(self.body + self.index * (not (self.format_index_main or self.format_index_light)))
        self.adjust_all_columns_width()
        
        # Save file
        self.workbook.save(filename=self.file_name)
        return

    # -------------------------------------------------------------------------
    # 2) Building blocks
    # -------------------------------------------------------------------------
    def correct_file_name(self, required_extension=".xlsx"):
        """
        Adds desired extention if the user did not include it in file name.

        ---------------
        Example:
        - File name without file extension: "Foo" --> "Foo.xlsx"
        """
        if not self.file_name.endswith(required_extension):
            self.file_name = self.file_name + required_extension
            print(f"The file name provided was modified to {self.file_name} to")
            print(" avoid errors in the program execution. To deactivate this ")
            print("autocorrection turn to 'False' the default arg 'check_file_name'.")
        return

    def check_file_existence(self):
        """
        Checks if in the directory exist a workbook and sheet under given names.

        ---------------
        It returns:
        - First bool: tells if a workbook called as the provided file name exists. 
        - Second bool: tells if inside the above workbook is present a sheet with the given name.
        - Workbook/None: if the workbook was found it returns it, else None.

        Note that for the sake of the complete program, if both the workbook and the
        sheet exists and the workbook has only one sheet, the function returns false
        and false. This is because, later, overwriting the only sheet will be
        equivalent to directly overwriting the whole workbook.
        """
        if path.exists(getcwd() + "/" + self.file_name):
            workbook = load_workbook(self.file_name)
            if self.sheet_name not in workbook.sheetnames:
                return [True, False], workbook
            else:
                if len(workbook.sheetnames) > 1:
                    return [True, True], workbook  # will overwrite just the sheet
                else:
                    return [False, False], None  # will replace directly the workbook
        else:
            return [False, False], None
        
    def save_df_to_excel(self):
        """
        Saves pandas dataframe to excel. Creates new notebooks or appends to old ones.

        ---------------
        Three outcomes:
        - If no workbook exists under the provided name:
            1) A new workbook is created
        - If it exists and if it already exists a sheet with the name given:
            2) The sheet/workbook gets overwritten
        - If it exists and no sheet with the name given exists:
            3) A new sheet is appended to the workbook
        """
        workbook_exist, sheet_exist = self.workbook_sheet_exist
        
        if not workbook_exist:
            self.df.to_excel(self.file_name, sheet_name=self.sheet_name,
                            index=self.keep_index)
        else:
            if sheet_exist: # it needs to be removed first
                name = self.workbook.get_sheet_by_name(self.sheet_name)
                self.workbook.remove_sheet(name)
                self.workbook.save(self.file_name)
            with pd.ExcelWriter(self.file_name, mode="a") as writer:
                self.df.to_excel(writer, sheet_name=self.sheet_name,
                                index=self.keep_index)
        return

    def find_index_columns_depth(self):
        """
        Returns number of multilevels for index and columns of a pandas dataframe.

        ---------------
        Takes a pandas dataframe and returns a list containing respectively the 
        number of levels of the multiindex and of the multicolumns. If the index 
        is simple the associated dimension is 1. Same is true for columns.
        """
        index_col_depth = []
        for dimension in [self.df.index, self.df.columns]:
            if isinstance(dimension, pd.MultiIndex):
                index_col_depth.append(len(dimension[0]))
            else:
                index_col_depth.append(1)
        return index_col_depth

    def find_header_coordinates(self):
        """
        Finds the top left cell of the header and the bottom right one.

        ---------------
        Returns a list containing two lists of length two: each contains a letter
        and a number, which define two cells in a spreadsheet.
        """
        starting_letter = string.ascii_uppercase[self.ind_depth * self.keep_index]
        starting_number = 1
        ending_letter = string.ascii_uppercase[self.ind_depth * self.keep_index + self.col_len - 1]
        ending_number = self.col_depth
        return [[starting_letter, starting_number], [ending_letter, ending_number]]

    def find_index_coordinates(self):
        """
        Finds the top left cell of the index and the bottom right one.

        ---------------
        Returns a list containing two lists of length two: each contains a letter
        and a number, which define two cells in a spreadsheet.
        """
        starting_letter = "A"
        starting_number = self.col_depth + 1
        ending_letter = string.ascii_uppercase[self.ind_depth - 1]
        ending_number = self.col_depth + self.ind_len
        return [[starting_letter, starting_number], [ending_letter, ending_number]]

    def rectangle_of_cells(self, coordinates_list):
        """
        Returns the cells belonging to a rectangular portion of a spreadsheet.

        ---------------
        Returns a list containing pairs in the form "A1". These correspond to the 
        cells contained in a rectangular portion of spreadsheet delimited by a top 
        left corner cell and a bottom right corner cell.
        
        Example:
        - If [[A,1],[B,2]] is provided, the output will be [A1, A2, B1, B2]
        """
        starting_letter, starting_number = coordinates_list[0]
        ending_letter, ending_number = coordinates_list[1]
        
        output_list = []
        increasing_letter, increasing_number = starting_letter, starting_number
        increasing_letter_position = string.ascii_uppercase.find(increasing_letter)
        ending_letter_position = string.ascii_uppercase.find(ending_letter)
        while increasing_letter_position <= ending_letter_position:
            while increasing_number <= ending_number:
                output_list.append(increasing_letter + str(increasing_number))
                increasing_number = increasing_number + 1
            increasing_number = starting_number
            increasing_letter_position = increasing_letter_position + 1
            increasing_letter = string.ascii_uppercase[increasing_letter_position]
        return output_list

    def get_workbook_sheet(self):
        """
        Returns the desired workbook and sheet objects based on their name.
        """
        workbook = load_workbook(self.file_name)
        sheet = workbook.get_sheet_by_name(self.sheet_name)
        return workbook, sheet
    
    def format_as_main(self, cells_list):
        """
        Applies formatting of the type "main" to a range of cells.
        
        ---------------
        The formatting style includes the font size, font boldness, font color,
        fill color, alignment.
        """
        main_font = Font(bold=self.main_font_bold, color=self.main_font_color,
                         size=self.main_font_size)
        main_fill = PatternFill(fill_type="solid", start_color=self.main_color,
                                end_color=self.main_color)
        main_alignment = Alignment(horizontal=self.main_alignment,
                                   vertical="center")
        self.apply_formatting_to_cells(cells_list, main_font, main_fill,
                                       main_alignment)
        return
    
    def format_as_light(self, cells_list):
        """
        Applies formatting of the type "light" to a range of cells.
        
        ---------------
        The formatting style includes the font size, font boldness, font color,
        fill color, alignment.
        """
        light_font = Font(bold=self.light_font_bold, color=self.light_font_color,
                          size=self.light_font_size)
        light_fill = PatternFill(fill_type="solid", start_color=self.light_color,
                                 end_color=self.light_color)
        light_alignment = Alignment(horizontal=self.light_alignment,
                                    vertical="center")
        self.apply_formatting_to_cells(cells_list, light_font, light_fill,
                                       light_alignment)
        return
    
    def format_body(self, cells_list):
        """
        Applies formatting to body cells.
        
        ---------------
        The formatting style includes the font size, font boldness, font color,
        fill color, alignment.
        """
        body_font = Font(size=self.body_font_size)
        body_alignment = Alignment(horizontal=self.body_alignment,
                                    vertical="center")
        self.apply_formatting_to_cells(cells_list, body_font, None,
                                       body_alignment)
        return

    def adjust_all_columns_width(self):
        """
        Sets the width of all the columns of the sheet to a fixed value
        """
        for value in range(0, self.col_len + 1):
            self.sheet.column_dimensions[string.ascii_uppercase[value]].width = self.custom_width
    # -------------------------------------------------------------------------
    # 3) Methods used at their times in building blocks 
    # -------------------------------------------------------------------------
    def apply_formatting_to_cells(self, cells_list, font_formatting=None,
                                  fill_formatting=None, alignment_formatting=None):
        """
        Applies a pre-determined formatting if any is provided
        """
        if font_formatting is not None:
            for cell in cells_list:
                self.sheet[cell].font = font_formatting
        if fill_formatting is not None:
            for cell in cells_list:
                self.sheet[cell].fill = fill_formatting
        if alignment_formatting is not None:
            for cell in cells_list:
                self.sheet[cell].alignment = alignment_formatting
        return

    # -------------------------------------------------------------------------
    # 4) Methods useful for debugging
    # -------------------------------------------------------------------------
    def print_dimensions(self):
        print("Header:", self.header)
        print()
        print("Index:", self.index)
        print()
        print("Body:", self.body)